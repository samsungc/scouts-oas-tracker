import io
from datetime import date, datetime

import openpyxl
from django.utils import timezone

from badges.models import BadgeRequirement
from submissions.models import BadgeSubmission
from users.models import User

BADGE_PREFIX_MAP = {
    "Aquatic": "Aquatic Skills",
    "Camping": "Camping Skills",
    "Chief Scout's Award": "Chief Scout's Award",
    "Emergency": "Emergency Skills",
    "North Star Award": "North Star Award",
    "Paddling": "Paddling Skills",
    "Sailing": "Sailing Skills",
    "Scoutcraft": "Scoutcraft Skills",
    "Seeonee Award": "Seeonee Award",
    "Trail": "Trail Skills",
    "Vertical": "Vertical Skills",
    "Winter": "Winter Skills",
}


def _make_aware(value):
    """Convert a date, naive datetime, or date string to an aware datetime."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                value = datetime.strptime(value, fmt)
                break
            except ValueError:
                continue
        else:
            return None
    if isinstance(value, date) and not isinstance(value, datetime):
        value = datetime(value.year, value.month, value.day)
    if timezone.is_naive(value):
        return timezone.make_aware(value)
    return value


def _parse_requirement_title(raw):
    """
    Convert 'Aquatic # 1.1' → 'Aquatic Skills 1.1'.
    Returns None if the format is unrecognised.
    """
    if not raw or "#" not in raw:
        return None
    parts = raw.split("#", 1)
    prefix = parts[0].strip()
    number = parts[1].strip()
    mapped = BADGE_PREFIX_MAP.get(prefix)
    if mapped is None:
        return None
    return f"{mapped} {number}"


def run_import(xlsx_file, dry_run=False):
    """
    Parse an xlsx badge-record spreadsheet and create approved BadgeSubmissions.

    Parameters
    ----------
    xlsx_file : file-like object or path string
        The xlsx file to parse.
    dry_run : bool
        If True, no database writes are performed.

    Returns
    -------
    dict with keys:
        processed, approved_rows, created, already_existed,
        warnings: { scout_not_found, requirement_not_found,
                    badge_prefix_not_found, reviewer_not_found }
    """
    # --- Build lookup tables (single DB round-trip each) ---
    scouts = {
        f"{u.first_name} {u.last_name}".lower(): u
        for u in User.objects.filter(role="scout")
        if u.first_name and u.last_name
    }

    requirements = {
        req.title.lower(): req
        for req in BadgeRequirement.objects.all()
    }

    # Reviewer lookup: username (case-insensitive) + "first.last" split
    reviewer_by_username = {
        u.username.lower(): u for u in User.objects.all()
    }
    reviewer_by_name = {
        f"{u.first_name}.{u.last_name}".lower(): u
        for u in User.objects.all()
        if u.first_name and u.last_name
    }

    existing = set(
        BadgeSubmission.objects
        .filter(status="approved")
        .values_list("scout_id", "requirement_id")
    )

    # --- Load workbook ---
    if isinstance(xlsx_file, (str, bytes)):
        wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
    else:
        content = xlsx_file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    sheet = wb.active  # badge_details_l2

    # Find the header row (look for the row containing "Candidate")
    header_row = None
    col_map = {}
    for row in sheet.iter_rows(max_row=10):
        for cell in row:
            if str(cell.value).strip().lower() == "candidate":
                header_row = cell.row
                break
        if header_row:
            break

    if header_row is None:
        return {"error": "Could not find header row in spreadsheet."}

    # Map column names → zero-based index
    for cell in sheet[header_row]:
        if cell.value:
            col_map[str(cell.value).strip().lower()] = cell.column - 1

    required_cols = {"candidate", "requirement", "submission date", "reviewed by", "reviewed date", "status"}
    missing = required_cols - set(col_map.keys())
    if missing:
        return {"error": f"Missing expected columns: {missing}"}

    # --- Iterate data rows ---
    warnings = {
        "scout_not_found": [],
        "requirement_not_found": [],
        "badge_prefix_not_found": [],
        "reviewer_not_found": [],
    }

    to_create = []
    processed = 0
    approved_rows = 0
    already_existed = 0

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        # Skip completely empty rows
        if not any(row):
            continue

        processed += 1

        status_val = row[col_map["status"]]
        if str(status_val).strip().lower() != "approved":
            continue

        approved_rows += 1

        candidate = str(row[col_map["candidate"]] or "").strip()
        req_raw = str(row[col_map["requirement"]] or "").strip()
        submission_date = row[col_map["submission date"]]
        reviewed_by_raw = str(row[col_map["reviewed by"]] or "").strip()
        reviewed_date = row[col_map["reviewed date"]]

        # Match scout
        scout = scouts.get(candidate.lower())
        if scout is None:
            if candidate and candidate not in warnings["scout_not_found"]:
                warnings["scout_not_found"].append(candidate)
            continue

        # Match requirement
        req_title = _parse_requirement_title(req_raw)
        if req_title is None:
            prefix = req_raw.split("#")[0].strip() if "#" in req_raw else req_raw
            if prefix and prefix not in warnings["badge_prefix_not_found"]:
                warnings["badge_prefix_not_found"].append(prefix)
            continue

        requirement = requirements.get(req_title.lower())
        if requirement is None:
            if req_title not in warnings["requirement_not_found"]:
                warnings["requirement_not_found"].append(req_title)
            continue

        # Skip if already approved
        if (scout.id, requirement.id) in existing:
            already_existed += 1
            continue

        # Match reviewer (optional — don't skip if missing)
        reviewer = None
        if reviewed_by_raw:
            reviewer = (
                reviewer_by_username.get(reviewed_by_raw.lower())
                or reviewer_by_name.get(reviewed_by_raw.lower())
            )
            if reviewer is None and reviewed_by_raw not in warnings["reviewer_not_found"]:
                warnings["reviewer_not_found"].append(reviewed_by_raw)

        submitted_at = _make_aware(submission_date)
        reviewed_at = _make_aware(reviewed_date)

        to_create.append(BadgeSubmission(
            scout=scout,
            requirement=requirement,
            status="approved",
            submitted_at=submitted_at,
            reviewed_at=reviewed_at,
            reviewed_by=reviewer,
        ))
        # Track in-memory so re-runs within the same call don't duplicate
        existing.add((scout.id, requirement.id))

    created = 0
    if not dry_run and to_create:
        BadgeSubmission.objects.bulk_create(to_create)
        created = len(to_create)
    elif dry_run:
        created = len(to_create)

    return {
        "processed": processed,
        "approved_rows": approved_rows,
        "created": created,
        "already_existed": already_existed,
        "dry_run": dry_run,
        "warnings": warnings,
    }
